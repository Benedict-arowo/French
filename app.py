from flask import Flask, request, render_template, send_file
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)


# Helper functions from your script
def load_attendance(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb[wb.sheetnames[0]]

    attendance = []
    phone_numbers_set = set()  # To track unique phone numbers

    for row_b in sheet.iter_rows(
        min_row=2, max_row=sheet.max_row, min_col=1, max_col=5
    ):
        phone_number = str(row_b[4].value).replace(" ", "")

        # Ensure phone number has correct length
        if len(phone_number) == 10:
            phone_number = "0" + phone_number

        # Skip if the phone number has already been added
        if phone_number in phone_numbers_set:
            continue  # This skips the current iteration and moves to the next row

        # Add the phone number to the set to track it
        phone_numbers_set.add(phone_number)

        full_name = row_b[2].value
        department = row_b[3].value

        attendance.append(
            {
                "full_name": full_name,
                "phone_number": phone_number,
                "department": department,
                "score": None,
            }
        )

    return attendance


def load_scores(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb[wb.sheetnames[0]]
    scores_dict = {}

    for row_b in sheet.iter_rows(
        min_row=2, max_row=sheet.max_row, min_col=1, max_col=5
    ):
        phone_number = str(row_b[3].value).replace(" ", "")
        score = row_b[2].value
        scores_dict[phone_number] = score
    return scores_dict


def process_data(attendance_file, scores_file):
    attendance = load_attendance(attendance_file)
    scores = load_scores(scores_file)

    final = []
    missed = []

    for student in attendance:
        if student["phone_number"] in scores:
            final.append(
                {
                    "full_name": student["full_name"],
                    "phone_number": student["phone_number"],
                    "score": scores[student["phone_number"]],
                }
            )
        elif str(student["phone_number"]).removeprefix("0") in scores:
            final.append(
                {
                    "full_name": student["full_name"],
                    "phone_number": student["phone_number"],
                    "score": scores[str(student["phone_number"]).removeprefix("0")],
                }
            )
        else:
            missed.append(
                {
                    "full_name": student["full_name"],
                    "phone_number": student["phone_number"],
                }
            )
    return final, missed


def save_results(final, missed, output_file="results.xlsx"):
    wb = Workbook()
    ws_final = wb.active
    ws_final.title = "Final Results"
    ws_final.append(["Full Name", "Phone Number", "Score"])

    for student in final:
        ws_final.append(
            [student["full_name"], student["phone_number"], student["score"]]
        )

    ws_missed = wb.create_sheet(title="Missed Students")
    ws_missed.append(["Full Name", "Phone Number"])

    for student in missed:
        ws_missed.append([student["full_name"], student["phone_number"]])

    wb.save(output_file)
    return output_file


# Flask Routes
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    if "attendance" not in request.files or "scores" not in request.files:
        return "Please upload both files", 400

    attendance_file = request.files["attendance"]
    scores_file = request.files["scores"]

    attendance_path = "uploaded_attendance.xlsx"
    scores_path = "uploaded_scores.xlsx"

    attendance_file.save(attendance_path)
    scores_file.save(scores_path)

    final, missed = process_data(attendance_path, scores_path)

    # Save results to a downloadable file
    result_file = save_results(final, missed)

    # Clean up uploaded files
    os.remove(attendance_path)
    os.remove(scores_path)

    return render_template(
        "results.html",
        final=sorted(final, key=lambda x: x["full_name"], reverse=False),
        missed=sorted(missed, key=lambda x: x["full_name"], reverse=False),
        result_file=result_file,
    )


@app.route("/download/<filename>")
def download(filename):
    return send_file(filename, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
