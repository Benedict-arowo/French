# Loop through all the numbers in sheet a
# While looping, check if the phone number exists in sheet b (sheet consisting the scores)
# If it does, insert the score into sheet a

from openpyxl import load_workbook


def load_attendance():
    wb = load_workbook(filename="engineering.xlsx")
    sheet = wb["Form Responses 1"]
    # for row in sheet:
    #     row[4].value = str(row[4].value).replace(" ", "")
    #     print(row[1].value, row[2].value, row[4].value)

    attendance = []

    for row_b in sheet.iter_rows(
        min_row=2, max_row=sheet.max_row, min_col=1, max_col=5
    ):
        phone_number = str(row_b[4].value).replace(" ", "")
        if len(phone_number) == 10:
            phone_number = "0" + phone_number
        full_name = row_b[2].value
        department = row_b[3].value

        attendance.append(
            {
                "full_name": full_name,
                "phone_number": phone_number,
                "department": department,
                "score": None,
            }
        )

    # print(attendance)
    return attendance


def load_scores():
    wb = load_workbook(filename="scores.xlsx")
    sheet = wb["Scores"]
    scores_dict = {}

    for row_b in sheet.iter_rows(
        min_row=2, max_row=sheet.max_row, min_col=1, max_col=5
    ):
        phone_number = str(row_b[3].value).replace(" ", "")
        score = row_b[2].value

        scores_dict[phone_number] = score

    # print(scores_dict)
    return scores_dict


def start():
    attedance = load_attendance()
    scores = load_scores()

    final = []
    missed = []

    for student in attedance:
        try:
            if scores[str(student["phone_number"])]:
                final.append(
                    {
                        "full_name": student["full_name"],
                        "phone_number": student["phone_number"],
                        "score": scores[str(student["phone_number"])],
                    }
                )
                # print(student, str(student["phone_number"]))
        except:
            missed.append(
                {
                    "full_name": student["full_name"],
                    "phone_number": student["phone_number"],
                }
            )
            pass

    print(final)
    # print(missed)


start()
